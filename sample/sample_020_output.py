import sys
import random
from datetime import datetime, timedelta

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QWidget, QPushButton, QFileDialog, QMessageBox
)
from PySide6.QtCharts import QChartView, QChart, QLineSeries, QDateTimeAxis, QValueAxis
from PySide6.QtCore import QDateTime, Qt, QPointF
from PySide6.QtGui import QPainter

# XLSXファイル書き込みのためのライブラリ
# pip install openpyxl が必要です
try:
    from openpyxl import Workbook
except ImportError:
    print("openpyxl ライブラリが見つかりません。'pip install openpyxl' を実行してください。")
    sys.exit(1)


class TrendChartApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ランダムデータによるトレンドチャート (PySide6 QChart)")
        self.setGeometry(100, 100, 1000, 700)  # x, y, width, height

        self.series = QLineSeries()
        self.series.setName("価格トレンド")

        num_data_points = 100
        start_time = datetime.now()
        price = 100.0

        for i in range(num_data_points):
            current_time = start_time + timedelta(minutes=i)
            timestamp_ms = int(current_time.timestamp() * 1000)

            price += random.uniform(-0.8, 0.8)
            price = max(50.0, min(150.0, price))

            self.series.append(QPointF(timestamp_ms, price))

        self.chart = QChart()
        self.chart.addSeries(self.series)
        self.chart.setTitle("株価トレンド")
        self.chart.legend().setVisible(True)
        self.chart.legend().setAlignment(Qt.AlignmentFlag.AlignBottom)

        axisX = QDateTimeAxis()
        axisX.setFormat("yyyy/MM/dd hh:mm:ss")
        axisX.setTitleText("時刻")
        self.chart.addAxis(axisX, Qt.AlignmentFlag.AlignBottom)
        self.series.attachAxis(axisX)

        axisY = QValueAxis()
        axisY.setLabelFormat("%.2f")
        axisY.setTitleText("価格")
        self.chart.addAxis(axisY, Qt.AlignmentFlag.AlignLeft)
        self.series.attachAxis(axisY)

        chart_view = QChartView(self.chart)
        chart_view.setRenderHint(QPainter.RenderHint.Antialiasing)

        save_button = QPushButton("XLSX形式でデータを保存")
        save_button.clicked.connect(self.save_data_to_xlsx)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        layout.addWidget(chart_view)
        layout.addWidget(save_button)

    def save_data_to_xlsx(self):
        """QLineSeriesから直接データを取得し、XLSXファイルとして保存するメソッド"""
        # ワークシート名を 'sample' に固定
        sheet_name = "sample"

        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "XLSXファイルを保存",
            "chart_book.xlsx",  # デフォルトのファイル名を変更
            "Excel Spreadsheet (*.xlsx);;All Files (*)"
        )

        if file_name:
            try:
                # 既存のブックを読み込むか、新しく作成する
                # 今回は常に新規ブックを作成
                wb = Workbook()

                # デフォルトで作成される'Sheet'という名前のシートを削除
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])

                # 固定された名前 'sample' で新しいシートを作成
                # openpyxlは、同じ名前のシートがある場合、自動的に (1), (2)などを付けて重複を避けます
                ws = wb.create_sheet(title=sheet_name, index=0)  # index=0 で先頭に作成

                # ヘッダー行を追加
                ws.append(["時刻 (ミリ秒タイムスタンプ)", "価格"])

                # QLineSeriesからデータポイントを取得し、書き込む
                for point in self.series.points():
                    timestamp_ms = int(point.x())
                    price_val = point.y()
                    ws.append([timestamp_ms, price_val])

                # ファイルを保存
                wb.save(file_name)
                QMessageBox.information(self, "保存完了",
                                        f"データを以下のファイルにワークシート '{sheet_name}' として保存しました:\n{file_name}")

            except Exception as e:
                QMessageBox.critical(self, "保存エラー", f"データの保存中にエラーが発生しました:\n{e}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = TrendChartApp()
    window.show()
    sys.exit(app.exec())
