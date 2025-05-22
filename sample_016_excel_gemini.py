import sys
import pandas as pd
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QToolBar, QToolButton,
    QStatusBar, QProgressBar, QFileDialog, QMessageBox,
    QStyle, QLabel  # QStyleとQLabelのインポートを追加
)
from PySide6.QtGui import QIcon, QAction
from PySide6.QtCore import (
    Qt, QThread, Signal, QObject
)
import openpyxl
import os


class ExcelLoader(QObject):
    # 進捗更新用のシグナル (現在のワークシート名, 現在のワークシート番号, 総ワークシート数)
    progress_updated = Signal(str, int, int)
    # 読み込み完了シグナル (読み込んだデータフレームの辞書)
    loading_finished = Signal(dict)
    # エラーシグナル
    error_occurred = Signal(str)

    def __init__(self, excel_path):
        super().__init__()
        self.excel_path = excel_path

    def run(self):
        try:
            # openpyxlは主に.xlsx, .xlsmをサポート。
            # .xls形式にはxlrdが必要だが、pandasのread_excelは自動で判断してくれるため、
            # openpyxlで対応できない場合はpandasのread_excelに任せるのが良い。
            # ただし、ここではシンプルにopenpyxlを使う前提で進める。
            # .xls形式の場合、openpyxl.load_workbookでエラーが出る可能性があるため、
            # その場合はpandas.read_excelを直接使うのが堅実。
            # 例外処理で対応するか、ファイル形式に応じて読み込みライブラリを切り替えるか。
            # 今回はプロンプトの「Excelは、ワークシートを調べて一つづ Pandasのデータフレームに読み込み」を優先し、
            # openpyxlで読み込み、DataFrameに変換する流れを維持。
            # openpyxlが対応しない.xlsの場合はエラーになることを許容する。

            # pandas.read_excelを使用すると、より簡単に読み込めるが、
            # プログレスバーの細かな更新（シートごと）は自前で制御する必要がある。
            # openpyxlでシートごとに読み込むことで、プログレスバーの更新を実現。

            workbook = openpyxl.load_workbook(self.excel_path)
            excel_data = {}
            total_sheets = len(workbook.sheetnames)

            for i, sheet_name in enumerate(workbook.sheetnames):
                self.progress_updated.emit(sheet_name, i + 1, total_sheets)
                sheet = workbook[sheet_name]
                data = sheet.values
                # 最初の行をヘッダーとしてPandasのDataFrameに読み込む
                # openpyxlから直接DataFrameを作成すると、空のセルがNoneになるため、適宜処理が必要な場合がある
                columns = next(data, None)  # ヘッダー行がない場合の対応
                if columns:
                    df = pd.DataFrame(data, columns=columns)
                else:  # ヘッダー行がない場合は、データ部分からDataFrameを作成
                    df = pd.DataFrame(data)

                excel_data[sheet_name] = df

            workbook.close()
            self.loading_finished.emit(excel_data)

        except Exception as e:
            self.error_occurred.emit(f"Excelファイルの読み込み中にエラーが発生しました: {e}\n"
                                     "'.xls'ファイルの場合、xlrdライブラリが不足しているか、\n"
                                     "openpyxlが対応していない可能性があります。")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excelファイルローダー (PySide6)")
        self.setGeometry(100, 100, 800, 600)

        self.excel_data_dict = {}  # 読み込んだExcelデータを保存する辞書

        self._create_toolbar()
        self._create_statusbar()

    def _create_toolbar(self):
        toolbar = QToolBar("ツールバー")
        self.addToolBar(toolbar)

        # SP_DirOpenIcon を表示したQToolButtonを作成
        open_action = QAction(self.style().standardIcon(QStyle.StandardPixmap.SP_DirOpenIcon), "Excelファイルを開く", self)
        open_action.triggered.connect(self._open_excel_file)
        toolbar.addAction(open_action)

    def _create_statusbar(self):
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)

        self.progressBar = QProgressBar(self)
        self.progressBar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progressBar.setTextVisible(True)
        self.progressBar.setMaximum(100)  # プログレスバーの最大値を100に設定
        self.progressBar.setValue(0)  # 初期値は0
        self.statusBar.addPermanentWidget(self.progressBar)  # 永続的に表示

        self.status_message_label = QLabel("準備完了", self)
        self.statusBar.addWidget(self.status_message_label)

    def _open_excel_file(self):
        # ファイルダイアログを開いてExcelファイルを選択
        file_dialog = QFileDialog(self)
        # フィルターに.xlsと.xlsmを追加
        file_dialog.setNameFilter("Excel Files (*.xlsx *.xls *.xlsm);;All Files (*)")
        file_dialog.setFileMode(QFileDialog.FileMode.ExistingFile)

        if file_dialog.exec():
            selected_files = file_dialog.selectedFiles()
            if selected_files:
                excel_path = selected_files[0]
                self.statusBar.showMessage(f"選択されたファイル: {os.path.basename(excel_path)}")
                self.progressBar.setValue(0)  # 新しいファイル選択時にプログレスバーをリセット
                self.status_message_label.setText("読み込み中...")
                self._load_excel_in_thread(excel_path)

    def _load_excel_in_thread(self, excel_path):
        # 既存のスレッドがあれば終了させる (複数回ファイルを開く場合を考慮)
        if hasattr(self, 'excel_thread') and self.excel_thread.isRunning():
            self.excel_thread.quit()
            self.excel_thread.wait()

        self.excel_thread = QThread()
        self.excel_loader = ExcelLoader(excel_path)
        self.excel_loader.moveToThread(self.excel_thread)

        # シグナルとスロットの接続
        self.excel_thread.started.connect(self.excel_loader.run)
        self.excel_loader.progress_updated.connect(self._update_progress_bar)
        self.excel_loader.loading_finished.connect(self._on_loading_finished)
        self.excel_loader.error_occurred.connect(self._on_error_occurred)
        self.excel_loader.loading_finished.connect(self.excel_thread.quit)  # 処理完了時にスレッドを終了
        self.excel_loader.error_occurred.connect(self.excel_thread.quit)  # エラー時にもスレッドを終了
        self.excel_thread.finished.connect(self.excel_thread.deleteLater)  # スレッドオブジェクトの削除

        # スレッドを開始
        self.excel_thread.start()

    def _update_progress_bar(self, sheet_name, current_sheet_num, total_sheets):
        progress_percentage = int((current_sheet_num / total_sheets) * 100)
        self.progressBar.setValue(progress_percentage)
        self.status_message_label.setText(f"読み込み中: シート '{sheet_name}' ({current_sheet_num}/{total_sheets})")

    def _on_loading_finished(self, excel_data_dict):
        self.excel_data_dict = excel_data_dict
        self.progressBar.setValue(100)
        self.statusBar.showMessage("Excelファイルの読み込みが完了しました！", 5000)  # 5秒間表示
        self.status_message_label.setText("読み込み完了")

        # 読み込んだデータの中身をターミナルに表示 (デバッグ用)
        QMessageBox.information(self, "読み込み完了",
                                "Excelファイルの読み込みが完了しました。\n詳細をコンソールに出力します。")
        for sheet_name, df in self.excel_data_dict.items():
            print(f"\n--- ワークシート: {sheet_name} ---")
            print(df.head())  # 各データフレームの先頭5行を表示
            print(f"行数: {len(df)}, 列数: {len(df.columns)}")

    def _on_error_occurred(self, error_message):
        self.statusBar.showMessage("エラー: Excelファイルの読み込みに失敗しました。", 5000)
        self.status_message_label.setText("エラー")
        QMessageBox.critical(self, "エラー", error_message)
        self.progressBar.setValue(0)  # エラー時はプログレスバーをリセット


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())