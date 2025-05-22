import openpyxl
import pandas as pd
from PySide6.QtCore import QObject, Signal

from funcs.tide import get_ymd
from structs.res import YMD


class ExcelLoader(QObject):
    # 進捗更新用のシグナル (現在のワークシート名, 現在のワークシート番号, 総ワークシート数)
    progressUpdated = Signal(str, int, int)
    # 読み込み完了シグナル (読み込んだデータフレームの辞書)
    finishedLoading = Signal(dict, YMD)
    # エラーシグナル
    errorOccurred = Signal(str)

    def __init__(self, excel_path: str):
        super().__init__()
        self.excel_path = excel_path
        # year: YYYY, month: MM, day: DD に分離
        self.ymd = get_ymd(excel_path)

    def run(self):
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            dict_sheet = dict()
            total_sheets = len(workbook.sheetnames)

            for i, name_sheet in enumerate(workbook.sheetnames):
                self.progressUpdated.emit(name_sheet, i + 1, total_sheets)
                sheet = workbook[name_sheet]
                data = sheet.values
                # 最初の行をヘッダーとしてPandasのDataFrameに読み込む
                # openpyxlから直接DataFrameを作成すると、空のセルがNoneになるため、適宜処理が必要な場合がある
                columns = next(data, None)  # ヘッダー行がない場合の対応
                if columns:
                    df = pd.DataFrame(data, columns=columns)
                else:  # ヘッダー行がない場合は、データ部分からDataFrameを作成
                    df = pd.DataFrame(data)
                dict_sheet[name_sheet] = df
            workbook.close()
            self.finishedLoading.emit(dict_sheet, self.ymd)

        except Exception as e:
            self.errorOccurred.emit(
                f"Excelファイルの読み込み中にエラーが発生しました: {e}\n"
                "'.xls'ファイルの場合、xlrdライブラリが不足しているか、\n"
                "openpyxlが対応していない可能性があります。"
            )
