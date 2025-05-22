import openpyxl
import pandas as pd
from PySide6.QtCore import QObject, Signal


class ExcelLoader(QObject):
    # 進捗更新用のシグナル (現在のワークシート名, 現在のワークシート番号, 総ワークシート数)
    progress_updated = Signal(str, int, int)
    # 読み込み完了シグナル (読み込んだデータフレームの辞書)
    finished_loading = Signal(dict)
    # エラーシグナル
    error_occurred = Signal(str)

    def __init__(self, excel_path: str):
        super().__init__()
        self.excel_path = excel_path

    def run(self):
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            excel_data = {}
            total_sheets = len(workbook.sheetnames)

            for i, sheet_name in enumerate(workbook.sheetnames):
                self.progress_updated.emit(sheet_name, i + 1, total_sheets)
                sheet = workbook[sheet_name]
                data = sheet.values
                # 最初の行をヘッダーとしてPandasのDataFrameに読み込む
                # openpyxlから直接DataFrameを作成すると、空のセルがNoneになるため、適宜処理が必要な場合がある
                columns = next(data, None)  # ヘッダー行がない場合の対応
                if columns:
                    df = pd.DataFrame(data, columns=columns)
                else:  # ヘッダー行がない場合は、データ部分からDataFrameを作成
                    df = pd.DataFrame(data)

                excel_data[sheet_name] = df

            workbook.close()
            self.finished_loading.emit(excel_data)

        except Exception as e:
            self.error_occurred.emit(f"Excelファイルの読み込み中にエラーが発生しました: {e}\n"
                                     "'.xls'ファイルの場合、xlrdライブラリが不足しているか、\n"
                                     "openpyxlが対応していない可能性があります。")
